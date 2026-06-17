from flask import Flask, request, jsonify, send_from_directory, abort
from flask_cors import CORS
import json
import os
import time
import threading
from datetime import datetime

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

DATA_DIR = 'data'
SURVEYS_FILE = os.path.join(DATA_DIR, 'surveys.json')
RESPONSES_FILE = os.path.join(DATA_DIR, 'responses.json')

lock = threading.Lock()

def ensure_data_dir():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

def load_json(filepath, default):
    ensure_data_dir()
    if not os.path.exists(filepath):
        return default
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return default

def save_json(filepath, data):
    ensure_data_dir()
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def generate_id():
    return str(int(time.time() * 1000)) + '_' + str(os.urandom(4).hex())

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/api/surveys', methods=['GET'])
def get_surveys():
    surveys = load_json(SURVEYS_FILE, [])
    return jsonify(surveys)

@app.route('/api/surveys/<survey_id>', methods=['GET'])
def get_survey(survey_id):
    surveys = load_json(SURVEYS_FILE, [])
    survey = next((s for s in surveys if s['id'] == survey_id), None)
    if not survey:
        return jsonify({'error': '问卷不存在'}), 404
    return jsonify(survey)

@app.route('/api/surveys', methods=['POST'])
def create_survey():
    data = request.json
    surveys = load_json(SURVEYS_FILE, [])
    
    survey = {
        'id': generate_id(),
        'title': data.get('title', '未命名问卷'),
        'description': data.get('description', ''),
        'questions': data.get('questions', []),
        'settings': {
            'startTime': data.get('settings', {}).get('startTime'),
            'endTime': data.get('settings', {}).get('endTime'),
            'maxResponses': data.get('settings', {}).get('maxResponses', 0),
            'password': data.get('settings', {}).get('password', ''),
            'allowModify': False
        },
        'createdAt': int(time.time() * 1000),
        'updatedAt': int(time.time() * 1000)
    }
    
    with lock:
        surveys.append(survey)
        save_json(SURVEYS_FILE, surveys)
    
    return jsonify(survey), 201

@app.route('/api/surveys/<survey_id>', methods=['PUT'])
def update_survey(survey_id):
    data = request.json
    surveys = load_json(SURVEYS_FILE, [])
    index = next((i for i, s in enumerate(surveys) if s['id'] == survey_id), None)
    
    if index is None:
        return jsonify({'error': '问卷不存在'}), 404
    
    with lock:
        surveys[index] = {
            **surveys[index],
            'title': data.get('title', surveys[index]['title']),
            'description': data.get('description', surveys[index].get('description', '')),
            'questions': data.get('questions', surveys[index]['questions']),
            'settings': {
                'startTime': data.get('settings', {}).get('startTime', surveys[index]['settings'].get('startTime')),
                'endTime': data.get('settings', {}).get('endTime', surveys[index]['settings'].get('endTime')),
                'maxResponses': data.get('settings', {}).get('maxResponses', surveys[index]['settings'].get('maxResponses', 0)),
                'password': data.get('settings', {}).get('password', surveys[index]['settings'].get('password', '')),
                'allowModify': False
            },
            'updatedAt': int(time.time() * 1000)
        }
        save_json(SURVEYS_FILE, surveys)
    
    return jsonify(surveys[index])

@app.route('/api/surveys/<survey_id>', methods=['DELETE'])
def delete_survey(survey_id):
    surveys = load_json(SURVEYS_FILE, [])
    responses = load_json(RESPONSES_FILE, [])
    
    with lock:
        surveys = [s for s in surveys if s['id'] != survey_id]
        responses = [r for r in responses if r['surveyId'] != survey_id]
        save_json(SURVEYS_FILE, surveys)
        save_json(RESPONSES_FILE, responses)
    
    return jsonify({'success': True})

@app.route('/api/surveys/<survey_id>/responses', methods=['GET'])
def get_survey_responses(survey_id):
    responses = load_json(RESPONSES_FILE, [])
    survey_responses = [r for r in responses if r['surveyId'] == survey_id]
    return jsonify(survey_responses)

@app.route('/api/surveys/<survey_id>/responses/count', methods=['GET'])
def get_response_count(survey_id):
    responses = load_json(RESPONSES_FILE, [])
    count = sum(1 for r in responses if r['surveyId'] == survey_id)
    return jsonify({'count': count})

@app.route('/api/surveys/<survey_id>/responses', methods=['POST'])
def submit_response(survey_id):
    data = request.json
    surveys = load_json(SURVEYS_FILE, [])
    survey = next((s for s in surveys if s['id'] == survey_id), None)
    
    if not survey:
        return jsonify({'success': False, 'error': '问卷不存在'}), 404
    
    now = int(time.time() * 1000)
    
    if survey['settings'].get('startTime'):
        start_time = int(datetime.fromisoformat(survey['settings']['startTime']).timestamp() * 1000)
        if now < start_time:
            return jsonify({'success': False, 'error': '问卷尚未开始'})
    
    if survey['settings'].get('endTime'):
        end_time = int(datetime.fromisoformat(survey['settings']['endTime']).timestamp() * 1000)
        if now > end_time:
            return jsonify({'success': False, 'error': '问卷已截止'})
    
    max_responses = survey['settings'].get('maxResponses', 0)
    if max_responses > 0:
        responses = load_json(RESPONSES_FILE, [])
        count = sum(1 for r in responses if r['surveyId'] == survey_id)
        if count >= max_responses:
            return jsonify({'success': False, 'error': '答题人数已达上限'})
    
    response = {
        'id': generate_id(),
        'surveyId': survey_id,
        'answers': data.get('answers', {}),
        'respondentInfo': data.get('respondentInfo', {}),
        'submittedAt': now
    }
    
    with lock:
        responses = load_json(RESPONSES_FILE, [])
        responses.append(response)
        save_json(RESPONSES_FILE, responses)
    
    return jsonify({'success': True, 'responseId': response['id']})

@app.route('/api/surveys/<survey_id>/status', methods=['GET'])
def get_survey_status(survey_id):
    surveys = load_json(SURVEYS_FILE, [])
    survey = next((s for s in surveys if s['id'] == survey_id), None)
    
    if not survey:
        return jsonify({'error': '问卷不存在'}), 404
    
    now = int(time.time() * 1000)
    responses = load_json(RESPONSES_FILE, [])
    count = sum(1 for r in responses if r['surveyId'] == survey_id)
    
    status = 'active'
    label = '进行中'
    
    if survey['settings'].get('endTime'):
        end_time = int(datetime.fromisoformat(survey['settings']['endTime']).timestamp() * 1000)
        if now > end_time:
            status = 'ended'
            label = '已结束'
    
    if survey['settings'].get('startTime'):
        start_time = int(datetime.fromisoformat(survey['settings']['startTime']).timestamp() * 1000)
        if now < start_time:
            status = 'pending'
            label = '未开始'
    
    if survey['settings'].get('maxResponses', 0) > 0 and count >= survey['settings']['maxResponses']:
        status = 'full'
        label = '已满员'
    
    return jsonify({'status': status, 'label': label})

@app.route('/api/init-sample', methods=['POST'])
def init_sample_data():
    surveys = load_json(SURVEYS_FILE, [])
    if len(surveys) > 0:
        return jsonify({'success': False, 'message': '已有数据，无需初始化'})
    
    q1_id = 'sample_q1'
    q2_id = 'sample_q2'
    q3_id = 'sample_q3'
    q4_id = 'sample_q4'
    q5_id = 'sample_q5'
    q6_id = 'sample_q6'
    
    sample_survey = {
        'id': 'sample_survey_001',
        'title': '用户满意度调查问卷',
        'description': '感谢您参与本次调查，您的反馈对我们非常重要！问卷大约需要3分钟完成。',
        'questions': [
            {
                'id': q1_id,
                'type': 'single',
                'title': '您的性别是？',
                'required': True,
                'options': ['男', '女', '其他'],
                'logic': [
                    {'operator': 'equals', 'value': '其他', 'action': 'show', 'target': q6_id}
                ]
            },
            {
                'id': q2_id,
                'type': 'single',
                'title': '您的年龄段是？',
                'required': True,
                'options': ['18岁以下', '18-25岁', '26-35岁', '36-45岁', '46岁以上'],
                'logic': []
            },
            {
                'id': q3_id,
                'type': 'multiple',
                'title': '您是通过哪些渠道了解到我们的产品的？（可多选）',
                'required': False,
                'options': ['朋友推荐', '社交媒体', '搜索引擎', '广告投放', '线下活动', '其他'],
                'logic': []
            },
            {
                'id': q4_id,
                'type': 'rating',
                'title': '请为我们的产品整体满意度打分',
                'required': True,
                'maxRating': 5,
                'logic': []
            },
            {
                'id': q5_id,
                'type': 'text',
                'title': '您对我们的产品有什么建议或意见？',
                'required': False,
                'placeholder': '请输入您的宝贵建议...',
                'logic': []
            },
            {
                'id': q6_id,
                'type': 'dropdown',
                'title': '您最常使用我们的哪项功能？',
                'required': False,
                'options': ['数据分析', '报表生成', '用户管理', '系统设置', '其他'],
                'logic': []
            }
        ],
        'settings': {
            'startTime': None,
            'endTime': None,
            'maxResponses': 0,
            'password': '',
            'allowModify': False
        },
        'createdAt': int(time.time() * 1000) - 86400000 * 7,
        'updatedAt': int(time.time() * 1000)
    }
    
    sample_responses = [
        {
            'id': 'sample_r1',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '男', q2_id: '26-35岁', q3_id: ['朋友推荐', '社交媒体'],
                q4_id: 5, q5_id: '产品很好用，界面也很美观。希望能增加更多自定义功能。', q6_id: '数据分析'
            },
            'submittedAt': int(time.time() * 1000) - 86400000 * 5
        },
        {
            'id': 'sample_r2',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '女', q2_id: '18-25岁', q3_id: ['搜索引擎', '广告投放'],
                q4_id: 4, q5_id: '整体不错，希望加载速度能更快一些。', q6_id: '报表生成'
            },
            'submittedAt': int(time.time() * 1000) - 86400000 * 4
        },
        {
            'id': 'sample_r3',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '男', q2_id: '36-45岁', q3_id: ['朋友推荐'],
                q4_id: 5, q5_id: '', q6_id: '用户管理'
            },
            'submittedAt': int(time.time() * 1000) - 86400000 * 3
        },
        {
            'id': 'sample_r4',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '女', q2_id: '26-35岁', q3_id: ['社交媒体', '线下活动'],
                q4_id: 3, q5_id: '功能还可以，但价格有点贵。', q6_id: '数据分析'
            },
            'submittedAt': int(time.time() * 1000) - 86400000 * 2
        },
        {
            'id': 'sample_r5',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '男', q2_id: '46岁以上', q3_id: ['广告投放'],
                q4_id: 4, q5_id: '界面简洁明了，容易上手。', q6_id: '系统设置'
            },
            'submittedAt': int(time.time() * 1000) - 86400000 * 1
        },
        {
            'id': 'sample_r6',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '女', q2_id: '18-25岁', q3_id: ['朋友推荐', '社交媒体', '搜索引擎'],
                q4_id: 5, q5_id: '非常喜欢！希望能推出更多新功能。', q6_id: '报表生成'
            },
            'submittedAt': int(time.time() * 1000) - 3600000 * 12
        },
        {
            'id': 'sample_r7',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '男', q2_id: '26-35岁', q3_id: ['其他'],
                q4_id: 2, q5_id: '有些功能不太好找，希望能优化一下导航。', q6_id: '其他'
            },
            'submittedAt': int(time.time() * 1000) - 3600000 * 6
        },
        {
            'id': 'sample_r8',
            'surveyId': 'sample_survey_001',
            'answers': {
                q1_id: '女', q2_id: '36-45岁', q3_id: ['朋友推荐', '广告投放'],
                q4_id: 4, q5_id: '', q6_id: '用户管理'
            },
            'submittedAt': int(time.time() * 1000) - 3600000 * 2
        }
    ]
    
    with lock:
        save_json(SURVEYS_FILE, [sample_survey])
        save_json(RESPONSES_FILE, sample_responses)
    
    return jsonify({'success': True, 'message': '示例数据已初始化'})

@app.route('/api/export/<survey_id>', methods=['GET'])
def export_survey_data(survey_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    surveys = load_json(SURVEYS_FILE, [])
    survey = next((s for s in surveys if s['id'] == survey_id), None)
    
    if not survey:
        return jsonify({'error': '问卷不存在'}), 404
    
    responses = load_json(RESPONSES_FILE, [])
    survey_responses = [r for r in responses if r['surveyId'] == survey_id]
    
    wb = Workbook()
    ws = wb.active
    ws.title = '答卷数据'
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    headers = ['序号', '提交时间'] + [q['title'] for q in survey['questions']]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    for row_idx, response in enumerate(survey_responses, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=datetime.fromtimestamp(response['submittedAt'] / 1000).strftime('%Y-%m-%d %H:%M:%S'))
        
        for col_idx, question in enumerate(survey['questions'], 3):
            answer = response['answers'].get(question['id'])
            if answer is None or answer == '':
                cell_value = ''
            elif isinstance(answer, list):
                cell_value = '、'.join(answer)
            elif isinstance(answer, dict) and question['type'] == 'matrix':
                cell_value = '；'.join([f'{k}:{v}' for k, v in answer.items()])
            elif question['type'] == 'rating':
                cell_value = f'{answer} 分'
            else:
                cell_value = str(answer)
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    filename = f"{survey['title']}_答卷数据.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    ensure_data_dir()
    app.run(host='0.0.0.0', port=8900, debug=True)
